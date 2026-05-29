import io
import logging

from aiogram import Bot, F, Router
from aiogram.fsm.context import FSMContext
from aiogram.fsm.state import State, StatesGroup
from aiogram.types import CallbackQuery, Message
from openpyxl import load_workbook

from src.config import SUPER_ADMIN_ID
from src.db import (
    add_admin,
    add_parcels_china,
    add_parcels_dushanbe,
    add_warehouse,
    delete_warehouse,
    find_in_china,
    find_in_dushanbe,
    get_parcels_by_client,
    get_setting,
    get_user_by_client_id,
    get_warehouse,
    list_admins,
    list_warehouses,
    mark_notified,
    remove_admin,
    set_setting,
    update_warehouse,
)
from src.fmt import (
    fmt_client_info_admin,
    fmt_parcel_arrived,
    fmt_track_result_admin,
    fmt_upload_result,
    fmt_warehouse_admin,
)
from src.keyboards import (
    admin_admins_kb,
    admin_back_kb,
    admin_main_kb,
    admin_wh_detail_kb,
    admin_wh_fields_kb,
    admin_warehouses_inline_kb,
    cancel_kb,
)

log = logging.getLogger(__name__)
router = Router(name="admin")


class AdminStates(StatesGroup):
    add_admin = State()
    remove_admin = State()
    check_track = State()
    check_client = State()
    upload_china = State()
    upload_dushanbe = State()
    wh_add_name = State()
    wh_add_phone = State()
    wh_add_region = State()
    wh_add_address = State()
    wh_edit_value = State()
    edit_tariffs = State()
    edit_support = State()


async def admin_start(msg: Message, state: FSMContext):
    await state.update_data(admin_mode=True)
    await state.set_state(None)
    await msg.answer(
        "🔧 Панель администратора",
        reply_markup=admin_main_kb(),
    )


# ── Навигация ──

@router.message(F.text == "❌ Отмена")
async def on_cancel(msg: Message, state: FSMContext):
    await state.set_state(None)
    await msg.answer(
        "🔧 Панель администратора",
        reply_markup=admin_main_kb(),
    )


@router.message(F.text == "⬅️ Назад в админку")
async def on_back_admin(
    msg: Message, state: FSMContext,
):
    await state.set_state(None)
    await msg.answer(
        "🔧 Панель администратора",
        reply_markup=admin_main_kb(),
    )


# ── Админы ──

@router.message(F.text == "👥 Админы")
async def on_admins_menu(
    msg: Message, state: FSMContext,
):
    await state.set_state(None)
    await msg.answer(
        "👥 Управление администраторами",
        reply_markup=admin_admins_kb(),
    )


@router.message(F.text == "📋 Список админов")
async def on_list_admins(msg: Message):
    admins = await list_admins()
    text = (
        "📋 Список админов:\n\n"
        f"👑 Суперадмин: {SUPER_ADMIN_ID}\n"
    )
    if admins:
        for a in admins:
            text += f"👤 {a}\n"
    else:
        text += "\nДополнительных админов нет."
    await msg.answer(
        text, reply_markup=admin_admins_kb(),
    )


@router.message(F.text == "➕ Добавить админа")
async def on_add_admin_start(
    msg: Message, state: FSMContext,
):
    await state.set_state(AdminStates.add_admin)
    await msg.answer(
        "Введите Telegram ID нового админа:",
        reply_markup=cancel_kb(),
    )


@router.message(AdminStates.add_admin)
async def on_add_admin_input(
    msg: Message, state: FSMContext,
):
    try:
        tid = int(msg.text.strip())
    except ValueError:
        await msg.answer(
            "⚠️ Введите числовой Telegram ID:",
            reply_markup=cancel_kb(),
        )
        return
    if await add_admin(tid):
        text = f"✅ Админ {tid} добавлен."
    else:
        text = f"⚠️ Админ {tid} уже существует."
    await state.set_state(None)
    await msg.answer(
        text, reply_markup=admin_admins_kb(),
    )


@router.message(F.text == "➖ Удалить админа")
async def on_remove_admin_start(
    msg: Message, state: FSMContext,
):
    await state.set_state(AdminStates.remove_admin)
    await msg.answer(
        "Введите Telegram ID для удаления:",
        reply_markup=cancel_kb(),
    )


@router.message(AdminStates.remove_admin)
async def on_remove_admin_input(
    msg: Message, state: FSMContext,
):
    try:
        tid = int(msg.text.strip())
    except ValueError:
        await msg.answer(
            "⚠️ Введите числовой Telegram ID:",
            reply_markup=cancel_kb(),
        )
        return
    if tid == SUPER_ADMIN_ID:
        text = "⚠️ Нельзя удалить суперадмина."
    elif await remove_admin(tid):
        text = f"✅ Админ {tid} удалён."
    else:
        text = f"⚠️ Админ {tid} не найден."
    await state.set_state(None)
    await msg.answer(
        text, reply_markup=admin_admins_kb(),
    )


# ── Загрузка Excel: Китай ──

@router.message(F.text == "📥 Загрузить Китай")
async def on_upload_china_start(
    msg: Message, state: FSMContext,
):
    await state.set_state(AdminStates.upload_china)
    await msg.answer(
        "📎 Отправьте Excel-файл.\n"
        "Столбец A: track_code",
        reply_markup=cancel_kb(),
    )


@router.message(
    AdminStates.upload_china, F.document,
)
async def on_upload_china_file(
    msg: Message, state: FSMContext, bot: Bot,
):
    doc = msg.document
    if not doc.file_name.endswith(
        (".xlsx", ".xls"),
    ):
        await msg.answer(
            "⚠️ Поддерживаются только .xlsx файлы.",
        )
        return
    file = await bot.get_file(doc.file_id)
    buf = io.BytesIO()
    await bot.download_file(file.file_path, buf)
    buf.seek(0)
    try:
        wb = load_workbook(buf, data_only=True)
        ws = wb.active
    except Exception as e:
        log.error("Ошибка чтения Excel: %s", e)
        await msg.answer(
            "⚠️ Не удалось прочитать файл.",
        )
        return
    tracks = []
    for row in range(2, ws.max_row + 1):
        val = ws.cell(row=row, column=1).value
        if val:
            tracks.append(str(val).strip())
    added = await add_parcels_china(tracks)
    await state.set_state(None)
    await msg.answer(
        fmt_upload_result(
            "ЗАГРУЗКА КИТАЙ", len(tracks), added,
        ),
        reply_markup=admin_main_kb(),
    )


# ── Загрузка Excel: Душанбе ──

@router.message(F.text == "📥 Загрузить Душанбе")
async def on_upload_dushanbe_start(
    msg: Message, state: FSMContext,
):
    await state.set_state(AdminStates.upload_dushanbe)
    await msg.answer(
        "📎 Отправьте Excel-файл.\n"
        "Столбец A: track_code\n"
        "Столбец B: client_id\n"
        "Столбец C: статус (+ или пусто)",
        reply_markup=cancel_kb(),
    )


@router.message(
    AdminStates.upload_dushanbe, F.document,
)
async def on_upload_dushanbe_file(
    msg: Message, state: FSMContext, bot: Bot,
):
    doc = msg.document
    if not doc.file_name.endswith(
        (".xlsx", ".xls"),
    ):
        await msg.answer(
            "⚠️ Поддерживаются только .xlsx файлы.",
        )
        return
    file = await bot.get_file(doc.file_id)
    buf = io.BytesIO()
    await bot.download_file(file.file_path, buf)
    buf.seek(0)
    try:
        wb = load_workbook(buf, data_only=True)
        ws = wb.active
    except Exception as e:
        log.error("Ошибка чтения Excel: %s", e)
        await msg.answer(
            "⚠️ Не удалось прочитать файл.",
        )
        return
    rows = []
    for row in range(2, ws.max_row + 1):
        track = ws.cell(row=row, column=1).value
        cid = ws.cell(row=row, column=2).value
        status_mark = (
            ws.cell(row=row, column=3).value or ""
        )
        if track and cid:
            rows.append((
                str(track).strip(),
                str(cid).strip(),
                str(status_mark).strip(),
            ))
    new_entries = await add_parcels_dushanbe(rows)
    await state.set_state(None)
    await msg.answer(
        fmt_upload_result(
            "ЗАГРУЗКА ДУШАНБЕ",
            len(rows), len(new_entries),
        ),
        reply_markup=admin_main_kb(),
    )
    # Уведомления только для новых waiting
    sent = 0
    for entry in new_entries:
        if entry["status"] == "received":
            continue
        user = await get_user_by_client_id(
            entry["client_id"],
        )
        if not user:
            continue
        lang = user.lang or "ru"
        try:
            await bot.send_message(
                chat_id=user.telegram_id,
                text=fmt_parcel_arrived(
                    entry["track_code"], lang,
                ),
            )
            await mark_notified(entry["track_code"])
            sent += 1
        except Exception as e:
            log.warning(
                "Не удалось отправить уведомление "
                "%s: %s",
                user.telegram_id, e,
            )
    if new_entries:
        await msg.answer(
            f"📨 Уведомлений отправлено: {sent}"
            f" из {len(new_entries)}",
            reply_markup=admin_main_kb(),
        )


# ── Проверки ──

@router.message(F.text == "🔎 Проверить трек")
async def on_check_track_start(
    msg: Message, state: FSMContext,
):
    await state.set_state(AdminStates.check_track)
    await msg.answer(
        "🔎 Введите трек-код:",
        reply_markup=cancel_kb(),
    )


@router.message(AdminStates.check_track)
async def on_check_track_input(
    msg: Message, state: FSMContext,
):
    text = msg.text.strip()
    in_china = await find_in_china(text)
    dushanbe = await find_in_dushanbe(text)
    user_info = None
    if dushanbe:
        user_info = await get_user_by_client_id(
            dushanbe.client_id,
        )
    await state.set_state(None)
    await msg.answer(
        fmt_track_result_admin(
            text.upper(), in_china,
            dushanbe, user_info,
        ),
        reply_markup=admin_back_kb(),
    )


@router.message(F.text == "🔎 Проверить клиента")
async def on_check_client_start(
    msg: Message, state: FSMContext,
):
    await state.set_state(AdminStates.check_client)
    await msg.answer(
        "🔎 Введите Client ID (TPS...):",
        reply_markup=cancel_kb(),
    )


@router.message(AdminStates.check_client)
async def on_check_client_input(
    msg: Message, state: FSMContext,
):
    user = await get_user_by_client_id(
        msg.text.strip(),
    )
    if not user:
        await state.set_state(None)
        await msg.answer(
            "❌ Клиент не найден.",
            reply_markup=admin_back_kb(),
        )
        return
    parcels = await get_parcels_by_client(
        user.client_id,
    )
    await state.set_state(None)
    await msg.answer(
        fmt_client_info_admin(user, parcels),
        reply_markup=admin_back_kb(),
    )


# ── Склады ──

FIELD_LABELS = {
    "name": "Название",
    "phone": "Телефон",
    "region": "Область",
    "address": "Адрес",
}


@router.message(F.text == "🏬 Склады")
async def on_warehouses_menu(
    msg: Message, state: FSMContext,
):
    await state.set_state(None)
    whs = await list_warehouses()
    await msg.answer(
        "🏬 Управление складами",
        reply_markup=admin_warehouses_inline_kb(whs),
    )


@router.callback_query(F.data == "awh_list")
async def on_awh_list(
    cb: CallbackQuery, state: FSMContext,
):
    await cb.answer()
    await state.set_state(None)
    whs = await list_warehouses()
    await cb.message.edit_text(
        "🏬 Управление складами",
        reply_markup=admin_warehouses_inline_kb(whs),
    )


@router.callback_query(
    F.data.startswith("awh_")
    & ~F.data.startswith("awhf_"),
)
async def on_awh_detail(
    cb: CallbackQuery, state: FSMContext,
):
    await cb.answer()
    data = cb.data

    if data == "awh_add":
        await state.set_state(
            AdminStates.wh_add_name,
        )
        await cb.message.answer(
            "Введите название склада:",
            reply_markup=cancel_kb(),
        )
        return

    if data.startswith("awh_del_"):
        wid = int(data.split("_")[2])
        if await delete_warehouse(wid):
            text = "✅ Склад удалён."
        else:
            text = "❌ Склад не найден."
        whs = await list_warehouses()
        await cb.message.edit_text(
            text + "\n\n🏬 Управление складами",
            reply_markup=admin_warehouses_inline_kb(
                whs,
            ),
        )
        return

    if data.startswith("awh_edit_"):
        wid = int(data.split("_")[2])
        w = await get_warehouse(wid)
        if not w:
            await cb.message.answer(
                "❌ Склад не найден.",
            )
            return
        await cb.message.edit_text(
            fmt_warehouse_admin(w)
            + "\n\nВыберите поле для изменения:",
            reply_markup=admin_wh_fields_kb(wid),
        )
        return

    wid = int(data.split("_")[1])
    w = await get_warehouse(wid)
    if not w:
        await cb.message.answer(
            "❌ Склад не найден.",
        )
        return
    await cb.message.edit_text(
        fmt_warehouse_admin(w),
        reply_markup=admin_wh_detail_kb(wid),
    )


@router.callback_query(F.data.startswith("awhf_"))
async def on_awh_field_select(
    cb: CallbackQuery, state: FSMContext,
):
    await cb.answer()
    parts = cb.data.split("_")
    wid = int(parts[1])
    field = parts[2]
    label = FIELD_LABELS.get(field, field)
    await state.update_data(
        wh_edit_id=wid, wh_edit_field=field,
    )
    await state.set_state(AdminStates.wh_edit_value)
    await cb.message.answer(
        f"Введите новое значение для «{label}»:",
        reply_markup=cancel_kb(),
    )


@router.message(AdminStates.wh_edit_value)
async def on_wh_edit_value(
    msg: Message, state: FSMContext,
):
    data = await state.get_data()
    wid = data["wh_edit_id"]
    field = data["wh_edit_field"]
    await update_warehouse(
        wid, field, msg.text.strip(),
    )
    w = await get_warehouse(wid)
    await state.set_state(None)
    await msg.answer(
        "✅ Обновлено.\n\n"
        + fmt_warehouse_admin(w),
        reply_markup=admin_wh_fields_kb(wid),
    )


@router.message(AdminStates.wh_add_name)
async def on_wh_add_name(
    msg: Message, state: FSMContext,
):
    await state.update_data(wh_name=msg.text.strip())
    await state.set_state(AdminStates.wh_add_phone)
    await msg.answer(
        "Введите телефон склада:",
        reply_markup=cancel_kb(),
    )


@router.message(AdminStates.wh_add_phone)
async def on_wh_add_phone(
    msg: Message, state: FSMContext,
):
    await state.update_data(wh_phone=msg.text.strip())
    await state.set_state(AdminStates.wh_add_region)
    await msg.answer(
        "Введите область/регион\n"
        "(напр. 新疆维吾尔自治区 "
        "乌鲁木齐市 天山区):",
        reply_markup=cancel_kb(),
    )


@router.message(AdminStates.wh_add_region)
async def on_wh_add_region(
    msg: Message, state: FSMContext,
):
    await state.update_data(
        wh_region=msg.text.strip(),
    )
    await state.set_state(AdminStates.wh_add_address)
    await msg.answer(
        "Введите адрес\n"
        "(напр. 延安路662号边疆宾馆19TPS号库房):",
        reply_markup=cancel_kb(),
    )


@router.message(AdminStates.wh_add_address)
async def on_wh_add_address(
    msg: Message, state: FSMContext,
):
    data = await state.get_data()
    wid = await add_warehouse(
        data["wh_name"], data["wh_phone"],
        data["wh_region"], msg.text.strip(),
    )
    w = await get_warehouse(wid)
    whs = await list_warehouses()
    await state.set_state(None)
    await msg.answer(
        "✅ Склад добавлен!\n\n"
        + fmt_warehouse_admin(w),
        reply_markup=admin_warehouses_inline_kb(whs),
    )


# ── Тарифы ──

@router.message(F.text == "💰 Тарифы")
async def on_edit_tariffs_start(
    msg: Message, state: FSMContext,
):
    current = (
        await get_setting("tariffs") or "(пусто)"
    )
    await state.set_state(AdminStates.edit_tariffs)
    await msg.answer(
        f"💰 Текущие тарифы:\n\n"
        f"{current}\n\n"
        "Отправьте новый текст для замены:",
        reply_markup=cancel_kb(),
    )


@router.message(AdminStates.edit_tariffs)
async def on_edit_tariffs_input(
    msg: Message, state: FSMContext,
):
    await set_setting("tariffs", msg.text.strip())
    await state.set_state(None)
    await msg.answer(
        "✅ Тарифы обновлены.",
        reply_markup=admin_main_kb(),
    )


# ── Поддержка ──

@router.message(F.text == "🆘 Поддержка")
async def on_edit_support_start(
    msg: Message, state: FSMContext,
):
    current = (
        await get_setting("support") or "(пусто)"
    )
    await state.set_state(AdminStates.edit_support)
    await msg.answer(
        f"🆘 Текущий текст поддержки:\n\n"
        f"{current}\n\n"
        "Отправьте новый текст для замены:",
        reply_markup=cancel_kb(),
    )


@router.message(AdminStates.edit_support)
async def on_edit_support_input(
    msg: Message, state: FSMContext,
):
    await set_setting("support", msg.text.strip())
    await state.set_state(None)
    await msg.answer(
        "✅ Поддержка обновлена.",
        reply_markup=admin_main_kb(),
    )